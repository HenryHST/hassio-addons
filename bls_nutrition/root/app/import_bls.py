"""Import BLS 4.0 XLSX files into SQLite."""

from __future__ import annotations

import re
import sqlite3
from pathlib import Path

from openpyxl import load_workbook

from app import db


def _parse_nutrient_header(header: str) -> tuple[str, str, str] | None:
    if not header or not isinstance(header, str):
        return None
    header = header.strip()
    if header.endswith("Datenherkunft"):
        return None
    if header.endswith("Referenz"):
        return None
    match = re.match(r"^([A-Z0-9]+)\s+(.+?)(?:\s+\[(.+)\])?$", header)
    if not match:
        return None
    code, name, unit = match.groups()
    return code, name.strip(), (unit or "").strip()


def import_components(conn: sqlite3.Connection, components_path: Path | None) -> None:
    if components_path is None or not components_path.exists():
        return

    workbook = load_workbook(components_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        workbook.close()
        return

    code_idx = None
    name_de_idx = None
    name_en_idx = None
    unit_idx = None
    group_idx = None
    for idx, header in enumerate(headers):
        label = str(header or "").lower()
        if "component code" in label or label == "code":
            code_idx = idx
        elif "deutsch" in label or "german" in label:
            name_de_idx = idx
        elif "english" in label:
            name_en_idx = idx
        elif "unit" in label:
            unit_idx = idx
        elif "group" in label:
            group_idx = idx

    if code_idx is None:
        workbook.close()
        return

    for row in rows:
        if not row or not row[code_idx]:
            continue
        conn.execute(
            """
            INSERT INTO nutrients(code, name_de, name_en, unit, nutrient_group)
            VALUES(?, ?, ?, ?, ?)
            ON CONFLICT(code) DO UPDATE SET
                name_de = excluded.name_de,
                name_en = excluded.name_en,
                unit = excluded.unit,
                nutrient_group = excluded.nutrient_group
            """,
            (
                str(row[code_idx]).strip(),
                str(row[name_de_idx]).strip() if name_de_idx is not None and row[name_de_idx] else None,
                str(row[name_en_idx]).strip() if name_en_idx is not None and row[name_en_idx] else None,
                str(row[unit_idx]).strip() if unit_idx is not None and row[unit_idx] else None,
                str(row[group_idx]).strip() if group_idx is not None and row[group_idx] else None,
            ),
        )
    workbook.close()


def _parse_value(raw: object) -> float | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    text = str(raw).strip()
    if not text or text in {"-", "TR", "tr"}:
        return None
    try:
        return float(text.replace(",", "."))
    except ValueError:
        return None


def import_data(conn: sqlite3.Connection, data_path: Path) -> int:
    workbook = load_workbook(data_path, read_only=True, data_only=True)
    sheet = workbook.active
    row_iter = sheet.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(row_iter)]

    nutrient_columns: list[tuple[int, str, str | None, str | None]] = []
    for idx, header in enumerate(headers):
        parsed = _parse_nutrient_header(header)
        if parsed:
            code, name, unit = parsed
            nutrient_columns.append((idx, code, name, unit))
            conn.execute(
                """
                INSERT INTO nutrients(code, name_de, unit)
                VALUES(?, ?, ?)
                ON CONFLICT(code) DO UPDATE SET
                    name_de = COALESCE(excluded.name_de, nutrients.name_de),
                    unit = COALESCE(excluded.unit, nutrients.unit)
                """,
                (code, name, unit or None),
            )

    imported = 0
    batch_foods: list[tuple[str, str, str | None]] = []
    batch_values: list[tuple[str, str, float | None, str | None, str | None]] = []

    for row in row_iter:
        if not row or not row[0]:
            continue
        bls_code = str(row[0]).strip()
        name_de = str(row[1]).strip() if len(row) > 1 and row[1] else bls_code
        name_en = str(row[2]).strip() if len(row) > 2 and row[2] else None
        batch_foods.append((bls_code, name_de, name_en))

        for idx, code, _, _ in nutrient_columns:
            if idx >= len(row):
                continue
            value = _parse_value(row[idx])
            provenance = None
            reference = None
            if idx + 1 < len(row):
                provenance = str(row[idx + 1]).strip() if row[idx + 1] is not None else None
            if idx + 2 < len(row):
                reference = str(row[idx + 2]).strip() if row[idx + 2] is not None else None
            if value is not None:
                batch_values.append((bls_code, code, value, provenance, reference))

        imported += 1
        if imported % 250 == 0:
            _flush_batches(conn, batch_foods, batch_values)
            batch_foods.clear()
            batch_values.clear()

    _flush_batches(conn, batch_foods, batch_values)
    workbook.close()
    return imported


def _flush_batches(
    conn: sqlite3.Connection,
    foods: list[tuple[str, str, str | None]],
    values: list[tuple[str, str, float | None, str | None, str | None]],
) -> None:
    if foods:
        conn.executemany(
            """
            INSERT INTO foods(bls_code, name_de, name_en)
            VALUES(?, ?, ?)
            ON CONFLICT(bls_code) DO UPDATE SET
                name_de = excluded.name_de,
                name_en = excluded.name_en
            """,
            foods,
        )
    if values:
        conn.executemany(
            """
            INSERT INTO food_nutrients(bls_code, nutrient_code, value, provenance, reference)
            VALUES(?, ?, ?, ?, ?)
            ON CONFLICT(bls_code, nutrient_code) DO UPDATE SET
                value = excluded.value,
                provenance = excluded.provenance,
                reference = excluded.reference
            """,
            values,
        )
